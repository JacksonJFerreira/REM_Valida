import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
import os
import openpyxl  # Para trabalhar diretamente com células em arquivos Excel
import logging

# Configuração do logging
logging.basicConfig(
    filename="script_logs.log",
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# Função para extrair informações de contrato e competência do REM
def extrair_informacoes_rem(caminho_arquivo):
    try:
        logging.info(f"Extraindo informações do arquivo REM: {caminho_arquivo}")
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        ws = wb.active

        contrato = ws['Q11'].value
        competencia = ws['Q7'].value

        if contrato is None or competencia is None:
            raise ValueError("As células Q11 ou Q7 estão vazias no arquivo REM.")

        logging.info(f"Contrato REM: {contrato}, Competência REM: {competencia}")
        return str(contrato), str(competencia)
    except Exception as e:
        logging.error(f"Erro ao extrair informações do arquivo REM: {caminho_arquivo}. Erro: {e}")
        return None, None

# Função para extrair informações de contrato e competência do SiFAC
def extrair_informacoes_sifac(caminho_arquivo):
    try:
        logging.info(f"Extraindo informações do arquivo SiFAC: {caminho_arquivo}")
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        ws = wb.active

        contrato = ws['C5'].value
        competencia = ws['F5'].value

        if contrato is None or competencia is None:
            raise ValueError("As células C5 ou F5 estão vazias no arquivo SiFAC.")

        contrato = contrato.replace("contrato:", "").strip()  # Remove o prefixo "contrato:"
        competencia = competencia.strip()

        logging.info(f"Contrato SiFAC: {contrato}, Competência SiFAC: {competencia}")
        return contrato, competencia
    except Exception as e:
        logging.error(f"Erro ao extrair informações do arquivo SiFAC: {caminho_arquivo}. Erro: {e}")
        return None, None

# Função para verificar e registrar os resultados nas planilhas
def verificar_e_registrar_planilhas():
    pasta_sifac = entrada_sifac.get()
    pasta_rem = entrada_rem.get()

    if not pasta_sifac or not pasta_rem:
        messagebox.showerror("Erro", "Por favor, selecione ambas as pastas.")
        logging.error("As pastas de entrada não foram selecionadas.")
        return

    arquivos_sifac = [os.path.join(pasta_sifac, f) for f in os.listdir(pasta_sifac) if f.endswith('.xlsx')]
    arquivos_rem = [os.path.join(pasta_rem, f) for f in os.listdir(pasta_rem) if f.endswith('.xlsx')]

    if not arquivos_sifac or not arquivos_rem:
        messagebox.showerror("Erro", "Não foram encontrados arquivos Excel em uma ou ambas as pastas.")
        logging.error("Não foram encontrados arquivos Excel em uma ou ambas as pastas.")
        return

    mensagem_resultado = ""

    for arquivo_rem in arquivos_rem:
        contrato_rem, competencia_rem = extrair_informacoes_rem(arquivo_rem)
        if not contrato_rem or not competencia_rem:
            mensagem_resultado += f"Erro ao extrair dados do arquivo REM: {arquivo_rem}\n"
            continue

        for arquivo_sifac in arquivos_sifac:
            contrato_sifac, competencia_sifac = extrair_informacoes_sifac(arquivo_sifac)
            if not contrato_sifac or not competencia_sifac:
                mensagem_resultado += f"Erro ao extrair dados do arquivo SiFAC: {arquivo_sifac}\n"
                continue

            # Validação de contrato e competência
            if contrato_rem == contrato_sifac and competencia_rem == competencia_sifac:
                mensagem_resultado += f"Contrato e competência correspondem entre: {arquivo_rem} e {arquivo_sifac}.\n"
                logging.info(f"Contrato e competência correspondem entre: {arquivo_rem} e {arquivo_sifac}.")

                # Comparar os nomes dos funcionários
                try:
                    rem_df = pd.read_excel(arquivo_rem, sheet_name="REM - Memória de Cálculo HHER", skiprows=8)
                    sifac_df = pd.read_excel(arquivo_sifac, skiprows=7)

                    nomes_rem = rem_df["D"].tolist()  # Coluna D dos arquivos REM
                    nomes_sifac = sifac_df["D"].tolist()  # Coluna D dos arquivos SiFAC

                    rem_df["Encontrado (SiFAC)"] = ["Sim" if nome in nomes_sifac else "Não" for nome in nomes_rem]

                    with pd.ExcelWriter(arquivo_rem, engine="openpyxl", mode="a") as writer:
                        rem_df.to_excel(writer, sheet_name="REM - Memória de Cálculo HHER", index=False, startrow=8)

                except Exception as e:
                    mensagem_resultado += f"Erro ao processar nomes entre: {arquivo_rem} e {arquivo_sifac}. Erro: {e}\n"
                    logging.error(f"Erro ao processar nomes: {e}")

            else:
                mensagem_resultado += f"Contrato ou competência não correspondem entre: {arquivo_rem} e {arquivo_sifac}.\n"
                logging.warning(f"Contrato ou competência não correspondem entre: {arquivo_rem} e {arquivo_sifac}.")

    resultado.set(mensagem_resultado)
    logging.info("Processamento concluído.")

# Função para selecionar a pasta
def selecionar_pasta(entry):
    pasta_selecionada = filedialog.askdirectory()
    entry.set(pasta_selecionada)
    logging.info(f"Pasta selecionada: {pasta_selecionada}")

# Interface gráfica
janela = ctk.CTk()
janela.title("Comparador de Planilhas com Logs")

entrada_sifac = ctk.StringVar()
entrada_rem = ctk.StringVar()
resultado = ctk.StringVar()

ctk.CTkLabel(janela, text="Selecione a pasta dos arquivos do SiFAC:").pack(pady=5)
ctk.CTkButton(janela, text="Selecionar Pasta", command=lambda: selecionar_pasta(entrada_sifac)).pack(pady=5)

ctk.CTkLabel(janela, text="Selecione a pasta dos arquivos do REM:").pack(pady=5)
ctk.CTkButton(janela, text="Selecionar Pasta", command=lambda: selecionar_pasta(entrada_rem)).pack(pady=5)

ctk.CTkButton(janela, text="Verificar e Registrar", command=verificar_e_registrar_planilhas).pack(pady=20)
ctk.CTkLabel(janela, textvariable=resultado, wraplength=500).pack(pady=10)

janela.mainloop()
